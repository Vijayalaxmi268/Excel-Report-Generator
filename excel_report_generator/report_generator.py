import pandas as pd
import matplotlib.pyplot as plt
from tkinter import Tk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import os

root = Tk()
root.withdraw()

file_path = filedialog.askopenfilename(
    title="Select CSV File",
    filetypes=[("CSV files", "*.csv")]
)
if not file_path:
    messagebox.showerror("Error", "No file selected!")
    exit()

df = pd.read_csv(file_path)

pivot = df.pivot_table(index="Region", columns="Product", values="Sales", aggfunc="sum", fill_value=0)

summary = df.describe(include='all')

plt.figure(figsize=(6, 4))
pivot.plot(kind='bar')
plt.title("Sales by Region & Product")
plt.ylabel("Sales")
chart_file = "sales_chart.png"
plt.savefig(chart_file)
plt.close()

wb = Workbook()
ws1 = wb.active
ws1.title = "Pivot Table"

for r in dataframe_to_rows(pivot, index=True, header=True):
    ws1.append(r)

for cell in ws1[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")

ws2 = wb.create_sheet("Summary Stats")
for r in dataframe_to_rows(summary, index=True, header=True):
    ws2.append(r)

from openpyxl.drawing.image import Image
img = Image(chart_file)
ws1.add_image(img, "H2")

save_path = filedialog.asksaveasfilename(
    title="Save Excel Report",
    defaultextension=".xlsx",
    filetypes=[("Excel Files", "*.xlsx")]
)
if save_path:
    wb.save(save_path)
    messagebox.showinfo("Success", f"Report saved to {save_path}")
else:
    messagebox.showerror("Error", "Save cancelled!")

if os.path.exists(chart_file):
    os.remove(chart_file)
